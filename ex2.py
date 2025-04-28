from openpyxl import Workbook, load_workbook

wb = Workbook()  # główny plik
ws = wb.active  # arkusz  z którym pracujemy

ws['A1'] = 42

ws.append([1, 2, 3])

wb.save('sample.xlsx')
wb.close()

print("Excel file has been created succesfully.")
print("file name: sample.xlsx")
print("location in current directory")

workbook = load_workbook('sample.xlsx')
sheet = workbook.active
print(sheet)  # <Worksheet "Sheet">

print(sheet['A1'].value)  # 42

print("###")
for row in sheet.iter_rows(min_row=1, max_row=3):
    for cell in row:
        print(cell.value)